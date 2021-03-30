#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''
参考文档: https://openpyxl.readthedocs.io/en/latest/
'''
import openpyxl       #pip3 install openpyxl
import os
import datetime


def demo_sheet(wb):
    print("------------------demo_sheet--------------------")
    ###get active(default create) worksheet
    ws = wb.active
    ws.title = "default-sheet"

    ###create new worksheet
    ws1 = wb.create_sheet("my-sheet")

    ###print sheet titles
    print(wb.sheetnames)     ###way 1
    for sheet in wb:         ###way 2
        print(sheet.title)

def demo_read(wb):
    print("------------------demo_read--------------------")
    try:
        wb["unexist"]        ###get sheet
    except:
        print("'wb[\"unexist\"]' except!!!")

    ws = wb["default-sheet"] ###get default sheet
    ws.cell(row=1, column=1, value=1)
    ws.cell(row=1, column=2, value=2)
    ws.cell(row=2, column=1, value=2)
    ws.cell(row=2, column=2, value=3)

    print("iter in rows")
    for r in ws.rows:
        for cell in r:
            print("\t", cell.value)

    print("iter in columns")
    for c in ws.columns:
        for cell in c:
            print("\t", cell.value)

    print("iter in values, by row")
    for r_vals in ws.values:
        print(r_vals)
        for val in r_vals:
            print("\t", val)

def demo_write(wb):
    print("------------------demo_write--------------------")
    ws = wb["my-sheet"]
    ws['A1'] = 42                        #write row1,column1
    ws['A2'] = datetime.datetime.now()   #write row2,column1
    
    ws.cell(row=1, column=2, value=12)   #write row1,column2
    cell = ws.cell(row=2, column=2)      #write row2,column2
    cell.value = 22

    for r_vals in ws.values:             #print values
        print(r_vals)


def demo_change(wb):
    print("------------------demo_change--------------------")
    wb.save("sample.xlsx")
    wb_new = openpyxl.load_workbook('sample.xlsx')

    print("after load/wb_new: ", wb_new.sheetnames)
    wb_new.create_sheet("change-sheet")
    print("after change/wb_new: ", wb_new.sheetnames)
    print("after change/wb: ", wb.sheetnames)
    os.system("rm sample.xlsx")
    
def demo():
    ###create workbook, default create a worksheet
    wb = openpyxl.Workbook()

    ###sub demo
    demo_sheet(wb)
    demo_change(wb)
    demo_read(wb)
    demo_write(wb)

if __name__ == "__main__":
    demo()
